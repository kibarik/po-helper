from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .deckspec import load_deckspec
from .theme import load_theme
from .build_slides import _collect_sprints


def _hexfill(theme_hex: str) -> PatternFill:
    h = theme_hex.lstrip("#").upper()
    return PatternFill("solid", fgColor=h)


def render_roadmap_file(spec_path, out_path, profile: dict | None = None) -> Path:
    spec = load_deckspec(spec_path)
    theme = load_theme((profile or {}).get("present"))
    sprints = _collect_sprints(spec)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Roadmap {spec.quarter}"[:31]
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # header row
    ws.cell(1, 1, "Инициативы в квартале").font = Font(bold=True)
    for j, sp in enumerate(sprints):
        c = ws.cell(1, 2 + j, sp)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _hexfill(theme.heading)
        c.alignment = wrap
    ws.column_dimensions["A"].width = 46
    for j in range(len(sprints)):
        ws.column_dimensions[chr(ord("B") + j)].width = 12

    r = 2
    for i, d in enumerate(spec.directions):
        fill = _hexfill(d.color or theme.direction_color(i))
        hc = ws.cell(r, 1, d.name)
        hc.font = Font(bold=True, color="FFFFFF")
        hc.fill = fill
        for j in range(len(sprints)):
            ws.cell(r, 2 + j).fill = _hexfill(d.color or theme.direction_color(i))
        r += 1
        for obj in d.objs:
            for kr in obj.krs:
                ws.cell(r, 1, kr.title).alignment = Alignment(wrap_text=True, vertical="center")
                for j, sp in enumerate(sprints):
                    cells = kr.sprint_cells.get(sp, [])
                    if cells:
                        cc = ws.cell(r, 2 + j, "\n".join(cells))
                        cc.alignment = wrap
                        cc.font = Font(size=9)
                r += 1

    r += 1
    ws.cell(r, 1, "Легенда фаз: Аналитика (BA/SA/ADR/PO) · Разработка (BE/FE) · "
                  "Отладка (QA) · Релиз (RM)").font = Font(italic=True, size=9)
    ws.cell(r + 1, 1, "S1 закоммичен из спринт-плана; далее — ориентировочно.").font = \
        Font(italic=True, size=9, color="8B94A6")

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out


if __name__ == "__main__":
    render_roadmap_file(sys.argv[1], sys.argv[2])
    print(f"wrote {sys.argv[2]}")
