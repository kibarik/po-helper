from pathlib import Path
from engine.render_roadmap import render_roadmap_file
from openpyxl import load_workbook

FIX = Path(__file__).parent / "fixtures" / "minimal_deck.yaml"

def test_roadmap_structure(tmp_path):
    out = render_roadmap_file(FIX, tmp_path / "roadmap.xlsx")
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert ws.title.startswith("Roadmap")
    values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert any("Надёжный источник" in v for v in values)     # direction header
    assert any("Продажа мероприятий" in v for v in values)   # KR row
    assert any("S1" in v for v in values)                     # sprint header
    assert any("SA·ADR" in v for v in values)                 # chip content
